"""Export functionality for generated layouts."""

import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from typing import Dict

from .config import (
    DNA_ORIGIN,
    DNA_DESTINATION,
    TRANSFECTION_DESTINATION,
    TRANSFECTION_TYPE,
    DNA_PART_NAME,
    DILUTED_SOURCE,
    PLATE_COLORS,
    OT2_SCRIPT_COLUMN_ORDER,
    get_layout,
)
from .utils import normalized_groupby


def generate_opentrons_script(
    df: pd.DataFrame,
    template_path: str = 'data/OT2_automated_transfection_v3.9.py',
    custom_template_content: str = None
) -> str:
    """
    Generate Opentrons Python script with embedded CSV config.

    Args:
        df: DataFrame with experiment configuration
        template_path: Path to template file (used if custom_template_content is None)
        custom_template_content: Custom template script content (overrides template_path)

    Returns:
        Opentrons script with CSV embedded
    """
    from .script_utils import prepare_script_for_export
    from .config import CIRCUIT_NAME, TRANSFECTION_GROUP

    # Make a copy to avoid modifying original
    df = df.copy()

    # Auto-generate Transfection type based on circuit/group DNA part count
    if CIRCUIT_NAME in df.columns and TRANSFECTION_GROUP in df.columns:
        with normalized_groupby(df, [CIRCUIT_NAME, TRANSFECTION_GROUP]) as grouped:
            df[TRANSFECTION_TYPE] = grouped[DNA_PART_NAME].transform(
                lambda x: 'Co' if len(x) > 1 else 'Single'
            )

    columns_to_export = [col for col in OT2_SCRIPT_COLUMN_ORDER if col in df.columns]
    csv_string = df[columns_to_export].to_csv(index=False)

    if custom_template_content:
        template = custom_template_content
    else:
        with open(template_path, 'r') as f:
            template = f.read()

    return prepare_script_for_export(template, csv_string)


def generate_excel_file(df: pd.DataFrame, plate_layouts: Dict[str, pd.DataFrame], layout_key: str = '24tube', labware_config: Dict[str, str] = None) -> BytesIO:
    """
    Generate Excel file with color-coded plate layouts.

    Args:
        df: Experiment design dataframe
        plate_layouts: Dictionary of layout DataFrames
        layout_key: Layout type ('24tube' or '96well')
        labware_config: Dictionary mapping slot numbers to labware types (unused, kept for compatibility)

    Returns:
        BytesIO containing Excel workbook
    """
    wb = Workbook()
    wb.remove(wb.active)

    layout = get_layout(layout_key)

    # Get reagent slots from static layout configuration
    reagent_slots = layout['reagent_slots']

    dna_parts = set(df[DNA_ORIGIN])
    destinations = set(df[DNA_DESTINATION]) | set(df[TRANSFECTION_DESTINATION])

    # Track diluted sources separately for distinct coloring
    diluted_sources_set = set()
    if DILUTED_SOURCE in df.columns:
        diluted_sources = df[DILUTED_SOURCE].dropna()
        diluted_sources_set.update(diluted_sources[diluted_sources != ''])

    for sheet_name, layout_df in plate_layouts.items():
        ws = wb.create_sheet(title=sheet_name)

        # Write headers
        ws.cell(1, 1, '')
        for col_idx, col_num in enumerate(layout_df.columns, start=2):
            cell = ws.cell(1, col_idx, str(col_num))
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write data
        for row_idx, row_label in enumerate(layout_df.index, start=2):
            cell = ws.cell(row_idx, 1, row_label)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            for col_idx, col_num in enumerate(layout_df.columns, start=2):
                value = layout_df.loc[row_label, col_num]
                cell = ws.cell(row_idx, col_idx, value if value else '')

                # Determine cell color
                if value:
                    if 'output_plate' in sheet_name:
                        fill_color = PLATE_COLORS['circuit']
                    else:
                        rack_num = sheet_name.split('_')[-1]
                        full_slot = f"{row_label}{col_num}.{rack_num}"

                        if full_slot in dna_parts:
                            fill_color = PLATE_COLORS['dna_part']
                        elif full_slot in diluted_sources_set:
                            fill_color = PLATE_COLORS['diluted_source']
                        elif full_slot in destinations:
                            fill_color = PLATE_COLORS['destination']
                        elif full_slot in reagent_slots:
                            fill_color = PLATE_COLORS['reagent']
                        else:
                            fill_color = PLATE_COLORS['empty']
                else:
                    fill_color = PLATE_COLORS['empty']

                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

        # Adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            ws.column_dimensions[col[0].column_letter].width = max(12, max_length + 2)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
